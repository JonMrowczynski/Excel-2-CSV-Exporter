"""
Copyright (c) 2018-2022 Jon Mrowczynski

Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated
documentation files (the "Software"), to deal in the Software without restriction, including without limitation the
rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to permit
persons to whom the Software is furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all copies or substantial portions of the
Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE
WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR
COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR
OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.

This script converts an Excel Workbook into one or more CSVs where each CSV file contains the contents of one Worksheet
of the Workbook. These CSVs are placed in a directory that is named after the Workbook. Each CSV is named after the
corresponding Worksheet.

This process can instead be done on a directory that contains multiple Excel Workbooks. For this case, everything is the
same except that the above process will be carried out on all of the Workbooks and the resulting directories will be
placed in a directory called "Exports".
"""

from argparse import ArgumentParser
from csv import writer
from os.path import isdir
from pathlib import Path
from typing import Final

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from tqdm import tqdm

WORKBOOK_EXTENSION: Final = 'xlsx'
EXPORT_ROOT_DIR: Final = 'Exports'


def _load_workbook(workbook_path: Path) -> Workbook:
    """
    Returns the loaded Workbook at the given path. If there does not exist a Workbook at the given path,
    then a FileNoteFoundError is raised. If the Workbook is open in another program, then a PermissionError is raised
    since only one program can exclusively access an Excel Workbook. Only the data is loaded from each of the
    Workbooks so that the values that are returned by formulas are exported instead of the actual formula.

    :param workbook_path: a relative or absolute path to a Workbook.
    :return: the loaded Workbook.
    """
    try:
        return load_workbook(workbook_path, data_only=True)
    except FileNotFoundError or PermissionError as e:
        problem_word = 'find' if isinstance(e, FileNotFoundError) else 'load'
        print(f'Could not {problem_word} {Workbook.__name__} "{workbook_path}".')


def _load_workbooks(root_dir_path: Path) -> dict[str: Workbook]:
    """
    Recursively scans for all Workbooks that can be found at the given root directory and loads them into a
    dictionary where the path to the Workbook is mapped to the loaded Workbook.

    :param root_dir_path: a relative or absolute path to a directory that contains Workbooks.
    :return: a dictionary containing paths to the Workbooks as the keys and the corresponding Workbooks as the values.
    """
    return {path: _load_workbook(path) for path in root_dir_path.rglob('*.' + WORKBOOK_EXTENSION)}


def _get_workbooks_map(input_path: Path) -> dict[Path, Workbook]:
    """
    Returns either a singleton dictionary if a path to a Workbook is given, a dictionary with multiple entries if a
    directory of Workbooks is given, or an empty dictionary if the given path is not valid or there are no Workbooks in
    the given dictionary.

    :param input_path: the path to the Workbook or directory of Workbooks that are to be converted.
    :return: a dictionary mapping Path(s) to Workbook(s) to the corresponding loaded Workbook(s).
    """
    if str(input_path).endswith(WORKBOOK_EXTENSION):
        print(f'Found input {Workbook.__name__} "{input_path}".')
        return {input_path: _load_workbook(input_path)}
    if isdir(input_path):
        print(f'Found input directory "{input_path}".')
        workbooks_map = _load_workbooks(input_path)
        if not workbooks_map:
            print(f'No {Workbook.__name__}s found in "{input_path}".')
        return workbooks_map
    print(f'"{input_path}" is not a valid {Workbook.__name__} or directory.')
    return dict()


def _should_write_data(export_path: Path) -> bool:
    """
    Determines whether data should be written to the given path. By default, data will be written unless there already
    exists a file at the given export path.

    :param export_path: the path that a CSV file will possible exported to.
    :return: a boolean indicating whether data should be written to the given export path.
    """
    if export_path.exists():
        print(f'Found "{export_path}".')
        choice = input('Would you like to overwrite? (y/[n]): ').lower()
        while choice != 'y' and choice != 'n':
            print('Please enter either "y" or "n"')
            choice = input('Would you like to overwrite? (y/[n]): ').lower()
        return True if choice == 'y' else False
    return True


def _remove_empty_rows(ws: Worksheet) -> Worksheet:
    """
    Removes all the rows in the given Worksheet that do not contain any data and returns this cleaned up Worksheet.

    :param ws: the Worksheet whose empty rows of data are to be deleted.
    :return: the cleaned up Worksheet
    """
    r = 1  # Excel starts indexing at 1.
    with tqdm(desc='Deleting empty rows', total=ws.max_row, position=0) as pbar:
        while r <= ws.max_row:
            if not any(cell.value for cell in ws[r]):
                ws.delete_rows(r)
            else:
                r += 1
            pbar.update()
    return ws


def _remove_empty_columns(ws: Worksheet) -> Worksheet:
    """
    Removes all the columns in the given Worksheet that do not contain any data and returns this cleaned up Worksheet.

    :param ws: the Worksheet whose empty columns of data are to be deleted.
    :return: the cleaned up Worksheet.
    """
    c = 1  # Excel starts indexing at 1.
    with tqdm(desc='Deleting empty columns', total=ws.max_column, position=0) as pbar:
        while c <= ws.max_column:
            if not any(row[0].value for row in ws.iter_rows(min_col=c, max_col=c)):
                ws.delete_cols(c)
            else:
                c += 1
            pbar.update()
    return ws


def _workbooks2csv(workbooks_map: dict[Path, Workbook], input_path: Path, output_path: Path) -> None:
    """
    Converts all the Worksheets in all the Workbooks into CSVs. One directory will be named after each Workbook and will
    contain all the data in the Workbook. Each CSV file will be named after each Worksheet and will contain the data
    present in the Worksheet.

    :param workbooks_map: maps Paths to Workbooks to their corresponding loaded Workbooks.
    :param input_path: the input Path that is a Workbook or is a directory containing Workbooks.
    """
    print(f'Converting {Workbook.__name__}s to CSVs...')
    for workbook_path, wb in workbooks_map.items():
        relative_workbook_path_without_extension = str(workbook_path.relative_to(input_path)).rpartition('.')[0]
        workbook_export_path = Path(output_path, relative_workbook_path_without_extension)
        workbook_export_path.mkdir(parents=True, exist_ok=True)
        for ws in wb:
            path2export = Path(workbook_export_path, ws.title + '.csv')
            if not _should_write_data(path2export):
                print(f'No data was written for {Worksheet.__name__} "{ws.title}".')
                continue
            with open(path2export, 'w', encoding='utf-8', newline='') as output_file:
                ws = _remove_empty_rows(_remove_empty_columns(ws))
                writer(output_file).writerows([[cell.value for cell in row] for row in ws])
                print(f'Successfully saved converted data to "{path2export}".')
    print(f'Converted {Workbook.__name__}s to CSVs!')


def parse_arguments() -> tuple[Path, Path]:
    parser = ArgumentParser()
    parser.add_argument('-input_path', type=str, required=True,
                        help=f'The path to the {Workbook.__name__} or to a directory containing {Workbook.__name__}s.')
    parser.add_argument('-output_path', type=str, required=False, default=EXPORT_ROOT_DIR, help='The path to the '
                                                                                                'directory that will '
                                                                                                'contain all the '
                                                                                                'exported CSV files.')
    args = parser.parse_args()
    input_path = Path(args.input_path)
    output_path = Path(args.output_path if args.output_path else EXPORT_ROOT_DIR)
    return input_path, output_path


def main():
    input_path, output_path = parse_arguments()
    if input_path == output_path:
        print('The input_path must be different than the output_path.')
    else:
        workbooks_map = _get_workbooks_map(input_path)
        if workbooks_map:
            _workbooks2csv(workbooks_map, input_path, output_path)


if __name__ == '__main__':
    main()
