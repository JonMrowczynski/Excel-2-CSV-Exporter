"""
This script allows one to convert an Excel Workbook into .csv files, where each .csv file orresponds to data in one
worksheet of the Excel Workbook. These .csv files are placed in a directory that has the same name as the name of the
Excel Workbook and the name of the .csv files are named after their corresponding worksheet.
"""

from typing import Union

from openpyxl import Workbook
from openpyxl.worksheet import Worksheet


def get_full_excel_workbook_name() -> Union[str, None]:
    """
    Finds every '.xlsx' Excel workbook in the directory that this script is in. If there are no Excel Workbooks found,
    then None is returned. Else if 1 Excel Workbook is found, then its full name is returned. Else more than one Excel
    Workbook was found, and the user is asked to pick one of the Excel Workbooks that they would like to be as the input
    Excel Workbook.

    :return: Either the only found Excel Workbook, the user chosen Excel Workbook, or None if no Excel Workbooks were
             found.
    """
    from os import listdir
    excel_workbook_names = [name for name in listdir('.') if name.endswith('.xlsx')]
    if not excel_workbook_names:
        print('Could not find any Excel Workbooks in current directory')
        print('Make sure that the Excel Workbook is in the same directory as this script')
        return None
    elif len(excel_workbook_names) == 1:
        print('Found Excel Workbook', excel_workbook_names[0])
        return excel_workbook_names[0]
    else:
        while True:
            print('More than one Excel Workbook was found in the current directory')
            print('Please choose one of the Excel Workbooks from the list')
            for i, name in enumerate(excel_workbook_names, 1):
                print(str(i) + '.', name)
            try:
                return excel_workbook_names[int(input('Choice: ')) - 1]
            except ValueError or IndexError:
                print('Please choose one of of the available options')


def load_workbook(full_workbook_name: str) -> Union[Workbook, None]:
    """
    Loads and returns the Excel Workbook that has the full name full_workbook_name. If the Excel Workbook could not be
    found, then None is returned.

    :param full_workbook_name: of the Excel Workbook that is to be loaded.
    :return: Either the Excel Workbook if it could be found, or None if it was not found.
    """
    if full_workbook_name:
        try:
            from openpyxl import load_workbook as load_excel_workbook
            return load_excel_workbook(full_workbook_name)
        except FileNotFoundError:
            print('Could not find Excel Workbook', full_workbook_name)
    return None


def get_output_file_relative_path(full_workbook_name: str, worksheet: Worksheet):
    """
    Returns the relative path to the .csv output file which will be located in a directory named after the Workbook and
    the name of the file will be named after the name given to the worksheet. In addition, the .csv file will also
    contain all of the information contained in the worksheet.

    :param full_workbook_name: that will be used to construct the directory name.
    :param worksheet: that will be used to construct the .csv file data and name.
    :return: the relative path to the .csv file.
    """
    from os.path import basename, splitext, dirname, sep, exists
    workbook_name = splitext(basename(full_workbook_name))[0]
    output_file_relative_path = workbook_name + sep + worksheet.title + '.csv'
    if exists(output_file_relative_path):
        print('Found', output_file_relative_path)
        choice = input('Would you like to overwrite? (y/n): ').lower()
        while choice != 'y' and choice != 'n':
            print('Please enter either "y" or "n"')
            choice = input('Would you like to overwrite? (y/n): ').lower()
        return output_file_relative_path if choice == 'y' else None
    elif not exists(dirname(output_file_relative_path)):
        from os import makedirs
        makedirs(dirname(output_file_relative_path))
    return output_file_relative_path


def convert_excel_to_csv(workbook: Workbook, full_workbook_name: str) -> None:
    """
    Converts a given Excel Workbook to .csv files, by outputting in a directory whose name is the name of the Excel
    Workbook, a .csv file that corresponds to each worksheet in the Excel Workbook whose file name corresponds to the
    name of the worksheet.

    :param workbook: that is to be converted to .csv files.
    :param full_workbook_name: of the Excel Workbook that is to be converted to .csv files.
    """
    for worksheet in workbook:
        output_file_relative_path = get_output_file_relative_path(full_workbook_name, worksheet)
        if output_file_relative_path:
            with open(output_file_relative_path, 'w', encoding='utf-8', newline='') as output_file:
                from csv import writer
                csv_writer = writer(output_file)
                for row in worksheet:
                    line = [cell.value for cell in row if cell.value is not None]
                    csv_writer.writerow(line)
                print('Data successfully saved to', output_file_relative_path)
        else:
            print('No data was written')


if __name__ == '__main__':
    full_workbook_name = get_full_excel_workbook_name()
    if full_workbook_name:
        workbook = load_workbook(full_workbook_name)
        if workbook:
            print('Successfully loaded Excel Workbook', full_workbook_name)
            convert_excel_to_csv(workbook, full_workbook_name)
    input('Press enter to exit...')
